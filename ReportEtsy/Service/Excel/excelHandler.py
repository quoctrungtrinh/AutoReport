from openpyxl import Workbook
from openpyxl import load_workbook
from Model.Constant import COLUMNS
import os
import shutil
import datetime

from Model.Constant import EXCEL_FILEPATH

class ExcelHandler:
    def __init__(self):                
        self.ColNames = COLUMNS
        
    def ColNamesToOrderAtts(self):
        atts = []
        for col in self.ColNames:
            if col == 'Year':
                atts.append('Year')
            if col == 'Month':
                atts.append('Month')
            if col == 'Date':
                atts.append('Date')
            if col == 'Order ID':
                atts.append('OrderID')
            if col == 'Article ID':
                atts.append('ArticleID')                
            if col == 'Transaction ID':
                atts.append('TransactionID')
            if col == 'Title':
                atts.append('Title')
            if col == 'Vendor':
                atts.append('Vendor')
            if col == 'Personalized':
                atts.append('isPersonalized')
            if col == 'Figure':
                atts.append('Figure')
            if col == 'Fulfillment Status':
                atts.append('FulfillmentStt')
            if col == 'Customer':
                atts.append('Kunden')
            if col == 'Address':
                atts.append('Address')
            if col == 'PLZ':
                atts.append('Plz')
            if col == 'City':
                atts.append('City')
            if col == 'Country':
                atts.append('Country')
            if col == 'Portal':
                atts.append('Portal')        
        return atts

    def Create_Report(self,wsTitle,rows):
        wb = Workbook()
        ws = wb.active

        ws.title = wsTitle
        ws.append(self.ColNames)

        for row in rows:            
            ws.append(row)
        

        wb.save(EXCEL_FILEPATH)

    def Append_Report(self,wsTitle,rows):
        wb = load_workbook(EXCEL_FILEPATH)
        ws = wb[wsTitle]

        for row in rows:
            ws.append(row)
        

        wb.save(EXCEL_FILEPATH)




        

        

